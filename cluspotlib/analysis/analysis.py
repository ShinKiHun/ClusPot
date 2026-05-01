"""
ClusterBench Analysis Module

Automatically detects MLIP dbs in work_dir/data/{mlip_name}/ and produces
analysis results under work_dir/analysis/

Directory structure:
  analysis/
    results.xlsx          ← all metrics (summary, per-model element data, FwT curves)
    {model}/
      {dataset}/          ← plots per dataset
        normal/
        anomaly/
      total/              ← aggregated plots for this model
    total/
      {dataset}/          ← comparison plots across all models
      total/              ← aggregated comparison plots

Execution order:
  Pass 1: Compute metrics and update xlsx (only for new/missing models)
  Pass 2: Generate per-model plots (only for models with incomplete PNGs)
  Pass 3: Generate total/global plots from xlsx (only if all models are complete)
"""

import gc
import math
import pickle
from collections import Counter
from pathlib import Path

import numpy as np
import pandas as pd
from ase.db import connect

from .metrics import energy_metrics, force_metrics, fwt_curve_data
from .plots import (
    plot_parity_elem,
    plot_periodic_heatmap,
    plot_total_bar,
    plot_pareto,
    plot_fwt_curve,
)

# Display names (with units) used in xlsx headers
HEADER_DISPLAY = {
    "model":                "Model",
    "dataset":              "Dataset",
    "element":              "Element",
    "type":                 "Type",
    "E_form_MAE":           "E_form MAE (eV/atom)",
    "E_form_RMSE":          "E_form RMSE (eV/atom)",
    "E_form_R2":            "E_form R²",
    "E_form_Pearson":       "E_form Pearson",
    "E_form_Spearman":      "E_form Spearman",
    "force_MAE":            "Force MAE (eV/Å)",
    "force_RMSE":           "Force RMSE (eV/Å)",
    "force_R2":             "Force R²",
    "force_Pearson":        "Force Pearson",
    "force_Spearman":       "Force Spearman",
    "force_cosine":         "Force cosine",
    "force_AFwT":           "AFwT",
    "n_normal":             "N_normal",
    "n_anomaly":            "N_anomaly",
    "n_missed":             "N_missed",
    "n_samples":            "N_samples",
    "anomaly_rate":         "Anomaly (%)",
    "cluster_time_med_s":   "Time_med (s)",
    "cluster_time_mean_s":  "Time_mean (s)",
    "cluster_time_total_s": "Time_total (s)",
}
# Reverse map: display name → raw key (first-occurrence wins for duplicates)
REVERSE_DISPLAY = {}
for _k, _v in HEADER_DISPLAY.items():
    if _v not in REVERSE_DISPLAY:
        REVERSE_DISPLAY[_v] = _k


def _normalize_xlsx_df(df: "pd.DataFrame") -> "pd.DataFrame":
    """Rename xlsx display-name columns back to raw keys."""
    return df.rename(columns=REVERSE_DISPLAY)


# Columns that should never be sanitized (counts, labels, times, fwt percentages)
SANITIZE_EXCLUDE = {
    "model", "dataset", "element", "type",
    "n_normal", "n_anomaly", "n_missed", "n_samples",
    "cluster_time_med_s", "cluster_time_mean_s", "cluster_time_total_s",
    "bulk_time_mean_s",
}


def _sanitize_value(val, header):
    """Replace missing or absurd metric values with the string '-'.

    Empty/None/NaN/Inf → '-'.
    R²-style columns: extreme negatives (< -10) → '-' (model effectively useless).
    Generic float metrics: |val| > 1e6 → '-'.
    Counts, labels, times, and fwt columns are returned untouched.
    """
    if header in SANITIZE_EXCLUDE:
        if val is None:
            return "-"
        return val
    h = str(header) if header is not None else ""
    # FwT@... columns are bounded percentages (0–100); only handle missing
    if h.startswith("FwT@"):
        if val is None:
            return "-"
        if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
            return "-"
        return val
    if val is None:
        return "-"
    if isinstance(val, float):
        if math.isnan(val) or math.isinf(val):
            return "-"
        if "R2" in h and val < -10:
            return "-"
        if abs(val) > 1e6:
            return "-"
    return val



def _is_single_element(element_label: str) -> bool:
    return "-" not in element_label


class ClusterAnalysis:
    """
    Analysis class for MLIP calculation results.

    Args:
        work_dir  : Working directory. Defaults to current directory (CWD).
        overwrite : If True, re-analyze already-analyzed entries (default: False)
        dpi       : Resolution for saved plots (default: 300)
    """

    def __init__(self, work_dir=None, overwrite: bool = False, dpi: int = 300):
        work_dir = Path(work_dir).resolve() if work_dir else Path.cwd()

        self.overwrite     = overwrite
        self.dpi           = dpi
        self.data_root     = work_dir / "data"
        self.analysis_root = work_dir / "analysis"
        self.xlsx_path     = self.analysis_root / "results.xlsx"

        self.analysis_root.mkdir(parents=True, exist_ok=True)

        # In-memory cache for DFT source data (shared across models within one run)
        self._dft_source_cache: dict = {}

    # ── Main execution ────────────────────────────────────────

    def run(self):
        src_db_map = {p.stem: p for p in sorted(self.data_root.glob("*.db"))}
        model_dirs = sorted(p for p in self.data_root.iterdir() if p.is_dir())

        if not model_dirs:
            print("No models found.")
            return

        print(f"Source DBs: {list(src_db_map.keys())}")
        print(f"Detected models: {[d.name for d in model_dirs]}")

        done_pairs = self._get_done_pairs()

        # ── Per-model loop: analyze + flush xlsx + plot, freeing memory between models ──
        # Each iteration handles one MLIP fully (Pass 1 metrics, xlsx flush, Pass 2 plots),
        # then releases its DataFrames so peak memory stays bounded to one model's data.
        for model_dir in model_dirs:
            mlip_name = model_dir.name

            # Per-model containers (scoped to this model only)
            new_summaries        : list = []
            model_total_summaries: list = []
            model_elem_data      = {mlip_name: {}}
            model_fwt_data       = {mlip_name: {}}
            per_dataset_dfs      : dict = {}    # dataset → (normal_df, anomaly_df)

            # ── Pass 1 (this model): compute metrics for each dataset ─────
            for result_db in sorted(model_dir.glob("*.db")):
                dataset_name = result_db.stem

                if dataset_name not in src_db_map:
                    print(f"[SKIP] {mlip_name}/{dataset_name}: source DB not found")
                    continue

                if (mlip_name, dataset_name) in done_pairs and not self.overwrite:
                    print(f"[xlsx OK] {mlip_name}/{dataset_name}: already in xlsx")
                    continue

                # Cache-first: rebuild metrics from parquet if available.
                cached_normal, cached_anomaly = self._load_df_cache(
                    mlip_name, dataset_name
                )
                if cached_normal is not None and len(cached_normal) > 0:
                    result = self._compute_metrics_from_dfs(
                        mlip_name, cached_normal, cached_anomaly, n_missed=0
                    )
                    source_label = "FromCache"
                else:
                    # No cache: read DFT source + MLIP DB.
                    dft_bulk, dft_cluster = self._load_dft_source(
                        dataset_name, src_db_map[dataset_name]
                    )
                    mlip_db = connect(str(result_db))
                    result = self._analyze_model(
                        mlip_name, mlip_db, dft_bulk, dft_cluster
                    )
                    source_label = "Computed"
                    if result[0] is not None:
                        # Persist cache so future runs can skip the DB read.
                        self._save_df_cache(
                            mlip_name, dataset_name, result[2], result[4]
                        )

                if result[0] is None:
                    print(f"[SKIP] {mlip_name}/{dataset_name}: no matching data")
                    continue

                (summary, normal_elem_df, normal_df,
                 anomaly_elem_df, anomaly_df,
                 thresholds, fwt_pct) = result
                summary["dataset"] = dataset_name
                new_summaries.append(summary)
                model_elem_data[mlip_name][dataset_name] = {
                    "normal":  normal_elem_df,
                    "anomaly": anomaly_elem_df,
                }
                if thresholds is not None:
                    model_fwt_data[mlip_name][dataset_name] = (thresholds, fwt_pct)
                per_dataset_dfs[dataset_name] = (normal_df, anomaly_df)

                print(f"[{source_label}] {mlip_name} / {dataset_name}")

            # ── Per-model total: combine new data with cached old datasets ─
            if per_dataset_dfs:
                all_model_data = dict(per_dataset_dfs)
                for result_db in sorted(model_dir.glob("*.db")):
                    ds_name = result_db.stem
                    if ds_name not in src_db_map or ds_name in all_model_data:
                        continue
                    cn, ca = self._load_df_cache(mlip_name, ds_name)
                    if cn is not None:
                        all_model_data[ds_name] = (cn, ca)

                total_result = self._build_model_total(mlip_name, all_model_data)
                if total_result is not None:
                    (total_summary, total_normal_elem, total_anomaly_elem,
                     _, _,
                     total_thresholds, total_fwt_pct) = total_result
                    model_total_summaries.append(total_summary)
                    model_elem_data[mlip_name]["total"] = {
                        "normal":  total_normal_elem,
                        "anomaly": total_anomaly_elem,
                    }
                    if total_thresholds is not None:
                        model_fwt_data[mlip_name]["total"] = (total_thresholds, total_fwt_pct)

                del all_model_data
                gc.collect()

            # ── Flush xlsx for this model only ─────────────────────────────
            # Each call merges with existing rows of OTHER models; only this
            # model's rows are added/replaced. So progress is preserved if a
            # later model fails.
            if new_summaries:
                self._update_xlsx(new_summaries, model_total_summaries,
                                  model_elem_data, model_fwt_data)

            # ── Pass 2 (this model): plots ────────────────────────────────
            for result_db in sorted(model_dir.glob("*.db")):
                dataset_name = result_db.stem
                if dataset_name not in src_db_map:
                    continue

                if self._has_complete_pngs(mlip_name, dataset_name) and not self.overwrite:
                    print(f"[SKIP] {mlip_name}/{dataset_name}: plots complete")
                    continue

                if dataset_name in per_dataset_dfs:
                    normal_df, anomaly_df = per_dataset_dfs[dataset_name]
                    normal_elem_df  = model_elem_data[mlip_name][dataset_name]["normal"]
                    anomaly_elem_df = model_elem_data[mlip_name][dataset_name]["anomaly"]
                else:
                    print(f"  [Re-read] {mlip_name}/{dataset_name}: PNG incomplete")
                    normal_df, anomaly_df = self._load_df_cache(mlip_name, dataset_name)
                    if normal_df is None:
                        print(f"  [Re-read DB] {mlip_name}/{dataset_name}: no cache, reading DB")
                        dft_bulk, dft_cluster = self._load_dft_source(
                            dataset_name, src_db_map[dataset_name]
                        )
                        mlip_db = connect(str(result_db))
                        (_, _, normal_df, _, anomaly_df, _, _) = self._analyze_model(
                            mlip_name, mlip_db, dft_bulk, dft_cluster
                        )
                        if normal_df is None:
                            print(f"  [SKIP] {mlip_name}/{dataset_name}: no data")
                            continue
                        self._save_df_cache(mlip_name, dataset_name, normal_df, anomaly_df)
                    normal_elem_df  = self._build_elem_df(normal_df)
                    anomaly_elem_df = self._build_elem_df(anomaly_df) \
                        if anomaly_df is not None and len(anomaly_df) > 0 else pd.DataFrame()

                print(f"\n[Plotting] {mlip_name} / {dataset_name}")
                dataset_dir = self.analysis_root / mlip_name / dataset_name
                dataset_dir.mkdir(parents=True, exist_ok=True)
                self._plot_subsets(mlip_name, normal_elem_df, normal_df,
                                   anomaly_elem_df, anomaly_df, dataset_dir)

            # Per-model total plot (regenerate when new data added)
            has_new_data = bool(per_dataset_dfs)
            if not self._has_complete_model_total(mlip_name) or self.overwrite or has_new_data:
                all_normal_dfs  = []
                all_anomaly_dfs = []
                for result_db in sorted(model_dir.glob("*.db")):
                    dataset_name = result_db.stem
                    if dataset_name not in src_db_map:
                        continue
                    if dataset_name in per_dataset_dfs:
                        nd, ad = per_dataset_dfs[dataset_name]
                    else:
                        nd, ad = self._load_df_cache(mlip_name, dataset_name)
                        if nd is None:
                            print(f"  [Re-read DB] {mlip_name}/{dataset_name}: no cache")
                            dft_bulk, dft_cluster = self._load_dft_source(
                                dataset_name, src_db_map[dataset_name]
                            )
                            mlip_db = connect(str(result_db))
                            (_, _, nd, _, ad, _, _) = self._analyze_model(
                                mlip_name, mlip_db, dft_bulk, dft_cluster
                            )
                            if nd is not None:
                                self._save_df_cache(
                                    mlip_name, dataset_name, nd,
                                    ad if ad is not None else pd.DataFrame()
                                )
                    if nd is not None and len(nd) > 0:
                        all_normal_dfs.append(nd)
                    if ad is not None and len(ad) > 0:
                        all_anomaly_dfs.append(ad)

                if all_normal_dfs:
                    combined_normal  = pd.concat(all_normal_dfs,  ignore_index=True)
                    combined_anomaly = pd.concat(all_anomaly_dfs, ignore_index=True) \
                        if all_anomaly_dfs else pd.DataFrame()
                    total_normal_elem  = self._build_elem_df(combined_normal)
                    total_anomaly_elem = self._build_elem_df(combined_anomaly) \
                        if len(combined_anomaly) > 0 else pd.DataFrame()
                    total_dir = self.analysis_root / mlip_name / "total"
                    total_dir.mkdir(parents=True, exist_ok=True)
                    self._plot_subsets(mlip_name, total_normal_elem, combined_normal,
                                       total_anomaly_elem, combined_anomaly, total_dir)
                    del combined_normal, combined_anomaly
                    del total_normal_elem, total_anomaly_elem
                del all_normal_dfs, all_anomaly_dfs

            # ── Free this model's memory before moving on ────────────────
            del new_summaries, model_total_summaries
            del model_elem_data, model_fwt_data, per_dataset_dfs
            # DFT source cache is keyed per-dataset and shared if same dataset
            # appears across models. Drop it here so it doesn't accumulate.
            self._dft_source_cache.clear()
            gc.collect()

        # ── Pass 3: total/global plots from xlsx ───────────────
        print("\n[Pass 3] Checking total plots...")
        all_complete = all(
            self._has_complete_pngs(model_dir.name, result_db.stem)
            for model_dir in model_dirs
            for result_db in sorted(model_dir.glob("*.db"))
            if result_db.stem in src_db_map
        )

        if all_complete or self.overwrite:
            print("[Pass 3] All models complete — generating total plots from xlsx...")
            self._plot_global_total_from_xlsx(list(src_db_map.keys()))
        else:
            print("[Pass 3] Some models still incomplete — skipping total plots.")

        print("\n[ALL DONE]")

    # ── PNG completeness checks ───────────────────────────────

    def _has_complete_pngs(self, mlip_name: str, dataset_name: str) -> bool:
        """
        Check if all required plot subdirectories exist and contain PNGs.
        normal/ is required.
        anomaly is optional (may not exist if no anomalies).
        """
        base = self.analysis_root / mlip_name / dataset_name
        d = base / "normal"
        if not d.exists() or not any(d.rglob("*.png")):
            return False
        return True

    def _has_complete_model_total(self, mlip_name: str) -> bool:
        """Check if per-model total plots are complete."""
        total_dir = self.analysis_root / mlip_name / "total"
        d = total_dir / "normal"
        if not d.exists() or not any(d.rglob("*.png")):
            return False
        return True

    # ── Plot subsets ──────────────────────────────────────────

    def _plot_subsets(self, mlip_name, normal_elem_df, normal_df,
                      anomaly_elem_df, anomaly_df, out_dir: Path):
        """Generate plots for normal/ and anomaly/."""
        normal_dir = out_dir / "normal"
        normal_dir.mkdir(parents=True, exist_ok=True)
        self._plot_model(mlip_name, normal_elem_df, normal_df, normal_dir)

        if anomaly_df is not None and len(anomaly_df) > 0:
            if anomaly_elem_df is not None and len(anomaly_elem_df) > 0:
                anomaly_dir = out_dir / "anomaly"
                anomaly_dir.mkdir(parents=True, exist_ok=True)
                self._plot_model(mlip_name, anomaly_elem_df, anomaly_df, anomaly_dir,
                                 skip_heatmap=True)

    # ── Build model total ─────────────────────────────────────

    def _build_model_total(self, mlip_name: str, dataset_dict: dict):
        """Aggregate data across all datasets. Returns data tuple without file I/O."""
        normal_dfs  = [v[0] for v in dataset_dict.values() if v[0] is not None and len(v[0]) > 0]
        anomaly_dfs = [v[1] for v in dataset_dict.values() if v[1] is not None and len(v[1]) > 0]

        if not normal_dfs:
            return None

        combined_normal  = pd.concat(normal_dfs,  ignore_index=True)
        combined_anomaly = pd.concat(anomaly_dfs, ignore_index=True) if anomaly_dfs else pd.DataFrame()

        normal_elem_df  = self._build_elem_df(combined_normal)
        anomaly_elem_df = self._build_elem_df(combined_anomaly) if len(combined_anomaly) > 0 else pd.DataFrame()

        thresholds, fwt_pct = fwt_curve_data(
            combined_normal["dft_forces"].tolist(),
            combined_normal["mlip_forces"].tolist(),
        )

        all_df = pd.concat([combined_normal, combined_anomaly], ignore_index=True) \
            if len(combined_anomaly) > 0 else combined_normal
        e_m = energy_metrics(combined_normal["e_form_dft"].values, combined_normal["e_form_mlip"].values)
        f_m = force_metrics(combined_normal["dft_forces"].tolist(), combined_normal["mlip_forces"].tolist())

        summary = {
            "model":                mlip_name,
            "dataset":              "total",
            "E_form_MAE":           e_m["MAE"],
            "E_form_RMSE":          e_m["RMSE"],
            "E_form_R2":            e_m["R2"],
            "E_form_Pearson":       e_m["Pearson"],
            "E_form_Spearman":      e_m["Spearman"],
            "force_MAE":            f_m["MAE"],
            "force_RMSE":           f_m["RMSE"],
            "force_R2":             f_m["R2"],
            "force_Pearson":        f_m["Pearson"],
            "force_Spearman":       f_m["Spearman"],
            "force_cosine":         f_m["cosine"],
            "force_AFwT":           f_m["AFwT"],
            "n_normal":             int(len(combined_normal)),
            "n_anomaly":            int(len(combined_anomaly)),
            "anomaly_rate":         float(len(combined_anomaly) / len(all_df) * 100),
            "n_missed":             0,
            "cluster_time_med_s":   float(np.nanmedian(combined_normal["calc_time"].values)),
            "cluster_time_mean_s":  float(np.nanmean(combined_normal["calc_time"].values)),
            "cluster_time_total_s": float(np.nansum(combined_normal["calc_time"].values)),
        }

        return (summary, normal_elem_df, anomaly_elem_df,
                combined_normal, combined_anomaly, thresholds, fwt_pct)

    # ── Filter helpers ────────────────────────────────────────

    # ── Loading ───────────────────────────────────────────────

    def _load_bulk_energy(self, db) -> dict:
        out = {}
        for row in db.select(type="bulk"):
            el = getattr(row, "element", None)
            if not el:
                continue
            e = getattr(row, "energy", None)
            n = getattr(row, "n_atoms", None)
            if e is None or n in (None, 0):
                continue
            out[el] = float(e) / float(n)
        return out

    def _load_cluster(self, db) -> dict:
        data = {}
        for row in db.select(type="cluster"):
            calc_id = getattr(row, "calc_id", None)
            step    = getattr(row, "step", None)
            if calc_id is None or step is None:
                continue
            atoms = row.toatoms()
            data[(calc_id, step)] = {
                "element": getattr(row, "element", None),
                "n_atoms": getattr(row, "n_atoms", None),
                "energy":  getattr(row, "energy", None),
                "forces":  atoms.get_forces().tolist(),
                "symbols": atoms.get_chemical_symbols(),
            }
        return data

    # ── DFT source cache ──────────────────────────────────────

    def _load_dft_source(self, dataset_name: str, src_db_path: Path):
        """Load DFT bulk + cluster data with disk & memory caching.

        Cache location:
          analysis/cache/{dataset_name}_dft_bulk.parquet
          analysis/cache/{dataset_name}_dft_cluster.parquet
        Falls back to legacy .pkl if parquet cache is missing.
        Cache is invalidated when the source DB is newer than the cache file.
        Within a single run, results are kept in memory so each dataset is
        loaded at most once regardless of how many models are analysed.
        """
        # 1) In-memory cache (within a single run)
        if dataset_name in self._dft_source_cache:
            return self._dft_source_cache[dataset_name]

        cache_dir    = self.analysis_root / "cache"
        bulk_path    = cache_dir / f"{dataset_name}_dft_bulk.parquet"
        cluster_path = cache_dir / f"{dataset_name}_dft_cluster.parquet"
        legacy_path  = cache_dir / f"{dataset_name}_dft.pkl"

        # 2) Parquet disk cache
        cache_file = bulk_path if bulk_path.exists() else legacy_path
        if cache_file.exists() and not self.overwrite:
            if cache_file.stat().st_mtime >= src_db_path.stat().st_mtime:
                print(f"  [Cache] {dataset_name}: loading DFT source from cache")
                if bulk_path.exists() and cluster_path.exists():
                    dft_bulk, dft_cluster = self._read_dft_parquet(bulk_path, cluster_path)
                else:
                    with open(legacy_path, "rb") as f:
                        data = pickle.load(f)
                    dft_bulk, dft_cluster = data["bulk"], data["cluster"]
                self._dft_source_cache[dataset_name] = (dft_bulk, dft_cluster)
                return dft_bulk, dft_cluster

        # 3) Cache miss: read from DB and save as parquet
        print(f"  [Cache] {dataset_name}: building DFT source cache (first time only)...")
        db          = connect(str(src_db_path))
        dft_bulk    = self._load_bulk_energy(db)
        dft_cluster = self._load_cluster(db)

        cache_dir.mkdir(parents=True, exist_ok=True)
        self._write_dft_parquet(dft_bulk, dft_cluster, bulk_path, cluster_path)
        print(f"  [Cache] {dataset_name}: cache saved → {bulk_path.parent}")

        self._dft_source_cache[dataset_name] = (dft_bulk, dft_cluster)
        return dft_bulk, dft_cluster

    def _write_dft_parquet(self, dft_bulk: dict, dft_cluster: dict,
                           bulk_path: Path, cluster_path: Path):
        """Serialize DFT source dicts to parquet files."""
        # bulk: simple element → energy_per_atom mapping
        bulk_df = pd.DataFrame(
            list(dft_bulk.items()), columns=["element", "energy_per_atom"]
        )
        bulk_df.to_parquet(bulk_path, index=False)

        # cluster: (calc_id, step) → {element, n_atoms, energy, forces, symbols}
        rows = [
            {
                "calc_id": calc_id,
                "step":    step,
                "element": v["element"],
                "n_atoms": v["n_atoms"],
                "energy":  v["energy"],
                "forces":  v["forces"],   # list[list[float]] — pyarrow handles this
                "symbols": v["symbols"],  # list[str]
            }
            for (calc_id, step), v in dft_cluster.items()
        ]
        cluster_df = pd.DataFrame(rows)
        cluster_df.to_parquet(cluster_path, index=False)

    def _read_dft_parquet(self, bulk_path: Path, cluster_path: Path):
        """Deserialize DFT source dicts from parquet files."""
        bulk_df  = pd.read_parquet(bulk_path)
        dft_bulk = dict(zip(bulk_df["element"], bulk_df["energy_per_atom"]))

        cluster_df  = pd.read_parquet(cluster_path)
        dft_cluster = {}
        for row in cluster_df.itertuples(index=False):
            # forces/symbols may be pyarrow arrays after parquet read — convert to list
            dft_cluster[(row.calc_id, row.step)] = {
                "element": row.element,
                "n_atoms": row.n_atoms,
                "energy":  row.energy,
                "forces":  [list(f) for f in row.forces],
                "symbols": list(row.symbols),
            }
        return dft_bulk, dft_cluster

    # ── Processed DataFrame cache ─────────────────────────────

    def _save_df_cache(self, mlip_name: str, dataset_name: str,
                       normal_df: pd.DataFrame, anomaly_df: pd.DataFrame):
        """Save processed normal/anomaly DataFrames to parquet cache."""
        cache_dir = self.analysis_root / mlip_name / dataset_name / "cache"
        cache_dir.mkdir(parents=True, exist_ok=True)
        normal_df.to_parquet(cache_dir / "normal_df.parquet",  index=False)
        anomaly_df.to_parquet(cache_dir / "anomaly_df.parquet", index=False)

    def _load_df_cache(self, mlip_name: str, dataset_name: str):
        """Load processed DataFrames from parquet cache.

        Falls back to legacy .pkl files if parquet cache is missing.
        Returns (normal_df, anomaly_df) or (None, None) if no cache found.
        """
        cache_dir    = self.analysis_root / mlip_name / dataset_name / "cache"
        normal_pq    = cache_dir / "normal_df.parquet"
        anomaly_pq   = cache_dir / "anomaly_df.parquet"
        normal_pkl   = cache_dir / "normal_df.pkl"
        anomaly_pkl  = cache_dir / "anomaly_df.pkl"

        if normal_pq.exists():
            normal_df  = pd.read_parquet(normal_pq)
            anomaly_df = pd.read_parquet(anomaly_pq) if anomaly_pq.exists() \
                         else pd.DataFrame()
            # pyarrow list<list<float>> columns may come back as non-native types;
            # explicitly convert to plain Python lists so np.asarray works downstream
            for col in ("dft_forces", "mlip_forces"):
                if col in normal_df.columns:
                    normal_df[col] = normal_df[col].apply(
                        lambda x: [list(f) for f in x] if x is not None else x
                    )
                if col in anomaly_df.columns:
                    anomaly_df[col] = anomaly_df[col].apply(
                        lambda x: [list(f) for f in x] if x is not None else x
                    )
            return normal_df, anomaly_df

        # Legacy fallback
        if normal_pkl.exists() and anomaly_pkl.exists():
            with open(normal_pkl,  "rb") as f:
                normal_df  = pickle.load(f)
            with open(anomaly_pkl, "rb") as f:
                anomaly_df = pickle.load(f)
            return normal_df, anomaly_df

        return None, None

    # ── Analysis ──────────────────────────────────────────────

    def _analyze_model(self, mlip_name, mlip_db, dft_bulk, dft_cluster):
        mlip_bulk = {}
        for row in mlip_db.select(type="bulk"):
            el = getattr(row, "element", None)
            if not el:
                continue
            epa = getattr(row, "energy_per_atom", None)
            if epa is None:
                e = getattr(row, "energy", None)
                n = getattr(row, "n_atoms", None)
                if e is None or n in (None, 0):
                    continue
                epa = float(e) / float(n)
            mlip_bulk[el] = float(epa)

        records = []
        miss    = 0

        for row in mlip_db.select(type="cluster"):
            calc_id = getattr(row, "calc_id", None)
            step    = getattr(row, "step", None)

            if calc_id is None or step is None:
                miss += 1
                continue

            dft = dft_cluster.get((calc_id, step), None)
            if dft is None:
                miss += 1
                continue

            elem    = dft.get("element", None)
            n       = dft.get("n_atoms", None)
            symbols = dft.get("symbols", None)
            if not elem or n in (None, 0) or not symbols:
                miss += 1
                continue

            comp = Counter(symbols)
            if not all(el in dft_bulk for el in comp):
                miss += 1
                continue
            if not all(el in mlip_bulk for el in comp):
                miss += 1
                continue

            dft_energy  = dft.get("energy", None)
            mlip_energy = getattr(row, "energy", None)
            if dft_energy is None or mlip_energy is None:
                miss += 1
                continue

            dft_ref  = sum(cnt * dft_bulk[el]  for el, cnt in comp.items())
            mlip_ref = sum(cnt * mlip_bulk[el] for el, cnt in comp.items())
            e_form_dft  = (float(dft_energy)  - dft_ref)  / float(n)
            e_form_mlip = (float(mlip_energy) - mlip_ref) / float(n)

            atoms_mlip  = row.toatoms()
            mlip_forces = atoms_mlip.get_forces().tolist()

            records.append({
                "element":     elem,
                "calc_id":     calc_id,
                "step":        int(step),
                "n_atoms":     int(n),
                "e_form_dft":  float(e_form_dft),
                "e_form_mlip": float(e_form_mlip),
                "dft_forces":  dft["forces"],
                "mlip_forces": mlip_forces,
                "calc_time":   float(getattr(row, "calc_time", np.nan)),
            })

        df = pd.DataFrame(records)
        if len(df) == 0:
            return None, None, None, None, None, None, None

        anomaly_mask = (df["e_form_dft"] - df["e_form_mlip"]).abs() >= 5.0
        normal_df    = df[~anomaly_mask].reset_index(drop=True)
        anomaly_df   = df[anomaly_mask].reset_index(drop=True)

        if len(normal_df) == 0:
            return None, None, None, None, None, None, None

        return self._compute_metrics_from_dfs(mlip_name, normal_df, anomaly_df, miss)

    def _compute_metrics_from_dfs(self, mlip_name, normal_df, anomaly_df, n_missed):
        """Aggregate metrics from already-split normal/anomaly DataFrames.

        Used by both _analyze_model (post-DB-read) and the Pass 1 cache-rebuild
        branch (when xlsx is missing but cache parquet exists).
        """
        if normal_df is None or len(normal_df) == 0:
            return None, None, None, None, None, None, None

        if anomaly_df is None:
            anomaly_df = pd.DataFrame()

        e_m = energy_metrics(normal_df["e_form_dft"].values, normal_df["e_form_mlip"].values)
        f_m = force_metrics(normal_df["dft_forces"].tolist(), normal_df["mlip_forces"].tolist())

        n_total = len(normal_df) + len(anomaly_df)
        summary = {
            "model":                mlip_name,
            "dataset":              None,
            "E_form_MAE":           e_m["MAE"],
            "E_form_RMSE":          e_m["RMSE"],
            "E_form_R2":            e_m["R2"],
            "E_form_Pearson":       e_m["Pearson"],
            "E_form_Spearman":      e_m["Spearman"],
            "force_MAE":            f_m["MAE"],
            "force_RMSE":           f_m["RMSE"],
            "force_R2":             f_m["R2"],
            "force_Pearson":        f_m["Pearson"],
            "force_Spearman":       f_m["Spearman"],
            "force_cosine":         f_m["cosine"],
            "force_AFwT":           f_m["AFwT"],
            "n_normal":             int(len(normal_df)),
            "n_anomaly":            int(len(anomaly_df)),
            "anomaly_rate":         float(len(anomaly_df) / n_total * 100) if n_total > 0 else 0.0,
            "n_missed":             int(n_missed),
            "cluster_time_med_s":   float(np.nanmedian(normal_df["calc_time"].values)),
            "cluster_time_mean_s":  float(np.nanmean(normal_df["calc_time"].values)),
            "cluster_time_total_s": float(np.nansum(normal_df["calc_time"].values)),
        }

        normal_elem_df  = self._build_elem_df(normal_df)
        anomaly_elem_df = self._build_elem_df(anomaly_df) if len(anomaly_df) > 0 else pd.DataFrame()

        thresholds, fwt_pct = fwt_curve_data(
            normal_df["dft_forces"].tolist(),
            normal_df["mlip_forces"].tolist(),
        )

        return summary, normal_elem_df, normal_df, anomaly_elem_df, anomaly_df, thresholds, fwt_pct

    def _build_elem_df(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame()
        rows = []
        for elem_g, grp in df.groupby("element"):
            e_em = energy_metrics(grp["e_form_dft"].values, grp["e_form_mlip"].values)
            f_em = force_metrics(grp["dft_forces"].tolist(), grp["mlip_forces"].tolist())
            rows.append({
                "element":         elem_g,
                "n_samples":       int(len(grp)),
                "E_form_MAE":      e_em["MAE"],
                "E_form_RMSE":     e_em["RMSE"],
                "E_form_R2":       e_em["R2"],
                "E_form_Pearson":  e_em["Pearson"],
                "E_form_Spearman": e_em["Spearman"],
                "force_MAE":       f_em["MAE"],
                "force_RMSE":      f_em["RMSE"],
                "force_R2":        f_em["R2"],
                "force_Pearson":   f_em["Pearson"],
                "force_Spearman":  f_em["Spearman"],
                "force_cosine":    f_em["cosine"],
                "cluster_time_mean_s": float(grp["calc_time"].mean()),
                "cluster_time_med_s":  float(grp["calc_time"].median()),
            })
        return pd.DataFrame(rows).sort_values("element").reset_index(drop=True)

    # ── xlsx ──────────────────────────────────────────────────

    def _get_done_pairs(self) -> set:
        """Determine done (model, dataset) pairs from xlsx per-model sheets.

        A pair is "done" only if all of the following hold:
          - the per-model element sheet has a row for that dataset
          - cache parquet (or legacy pkl) exists on disk
          - the summary sheet has a row for that (model, dataset)

        Legacy xlsx without a "dataset" column in summary fails the third
        check on every pair, so the cache-first path will backfill per-
        dataset summary rows on the next run.
        """
        done = set()
        if not self.xlsx_path.exists():
            return done
        try:
            xl = pd.ExcelFile(self.xlsx_path)

            summary_pairs   = set()
            summary_has_col = False
            if "summary" in xl.sheet_names:
                sum_df = _normalize_xlsx_df(pd.read_excel(xl, sheet_name="summary"))
                if "model" in sum_df.columns and "dataset" in sum_df.columns:
                    summary_has_col = True
                    sum_df["model"] = sum_df["model"].ffill()
                    for _, r in sum_df.iterrows():
                        m  = r.get("model")
                        ds = r.get("dataset")
                        if pd.isna(m) or pd.isna(ds):
                            continue
                        if str(ds) != "total":
                            summary_pairs.add((str(m), str(ds)))

            for sheet in xl.sheet_names:
                if sheet in ("summary", "fwt_curves"):
                    continue
                df = pd.read_excel(xl, sheet_name=sheet)
                # Column may be stored as "Dataset" (display name) or "dataset" (raw)
                dataset_col = next(
                    (c for c in df.columns if c.lower() == "dataset"), None
                )
                if dataset_col is None:
                    continue
                for ds in df[dataset_col].dropna().unique():
                    if str(ds) != "total":
                        cache_dir = self.analysis_root / sheet / str(ds) / "cache"
                        cache_ok = (
                            (cache_dir / "normal_df.parquet").exists() or
                            (cache_dir / "normal_df.pkl").exists()
                        )
                        summary_ok = (
                            summary_has_col and (sheet, str(ds)) in summary_pairs
                        )
                        if cache_ok and summary_ok:
                            done.add((sheet, str(ds)))
        except Exception:
            pass
        return done

    def _update_xlsx(self, new_summaries, model_total_summaries,
                     model_elem_data, model_fwt_data):
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        EXCLUDE_COLS = {"bulk_time_mean_s"}
        FMT_INT   = "#,##0"
        FMT_PCT   = "0.00"
        FMT_FLOAT = "0.000"
        INT_COLS  = {"n_normal", "n_anomaly", "n_missed", "n_samples"}
        PCT_COLS  = {"anomaly_rate"}

        def _autosize_columns(ws, headers, min_w=10, max_w=45, pad=2):
            """Set each column width to fit header + longest cell content."""
            for col_idx, h in enumerate(headers, 1):
                col_letter = get_column_letter(col_idx)
                display = HEADER_DISPLAY.get(h, h)
                max_len = len(str(display))
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is None:
                            continue
                        if isinstance(cell.value, float):
                            s = f"{cell.value:.4f}"
                        else:
                            s = str(cell.value)
                        if len(s) > max_len:
                            max_len = len(s)
                ws.column_dimensions[col_letter].width = min(
                    max(max_len + pad, min_w), max_w
                )

        def _apply_sheet_style(ws, headers):
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value     = HEADER_DISPLAY.get(h, h)
                cell.font      = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            ws.row_dimensions[1].height = 20

            _autosize_columns(ws, headers)

            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                ws.row_dimensions[row_idx].height = 25
                for col_idx, cell in enumerate(row, 1):
                    h = headers[col_idx - 1]
                    is_label = (col_idx == 1) or (h in ("model", "element", "dataset", "type") and col_idx <= 3)
                    cell.font      = Font(bold=is_label, size=11)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Replace missing or absurd metric values with "-"
                    sanitized = _sanitize_value(cell.value, h)
                    if sanitized == "-" and cell.value != "-":
                        cell.value = "-"
                        cell.number_format = "@"
                        continue

                    if h in INT_COLS:
                        cell.number_format = FMT_INT
                    elif h in PCT_COLS:
                        cell.number_format = FMT_PCT
                    elif isinstance(cell.value, float):
                        cell.number_format = FMT_FLOAT

        def _merge_first_col(ws):
            """Merge consecutive cells with the same value in column 1 (skip header row)."""
            max_row = ws.max_row
            if max_row < 3:
                return
            group_start = 2
            current_val = ws.cell(row=2, column=1).value
            for row_idx in range(3, max_row + 2):
                val = ws.cell(row=row_idx, column=1).value if row_idx <= max_row else object()
                if val != current_val:
                    if row_idx - 1 > group_start:
                        ws.merge_cells(start_row=group_start, start_column=1,
                                       end_row=row_idx - 1, end_column=1)
                    ws.cell(row=group_start, column=1).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    current_val = val
                    group_start = row_idx

        def _apply_borders(ws, group_col=1):
            """Apply thin grid borders, medium header separator, and medium group separators."""
            thin   = Side(style='thin')
            medium = Side(style='medium')

            max_row = ws.max_row
            max_col = ws.max_column
            if max_row < 1 or max_col < 1:
                return

            # Thin grid on all cells
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).border = Border(
                        left=thin, right=thin, top=thin, bottom=thin
                    )

            # Medium bottom on header row
            for col_idx in range(1, max_col + 1):
                c = ws.cell(row=1, column=col_idx)
                c.border = Border(left=thin, right=thin, top=thin, bottom=medium)

            if group_col is None or max_row < 3:
                return

            # Find group starts: rows where group_col has non-None value (top of each merged group)
            group_starts = [r for r in range(2, max_row + 1)
                            if ws.cell(row=r, column=group_col).value is not None]
            if len(group_starts) < 2:
                return
            group_ends = [group_starts[i + 1] - 1 for i in range(len(group_starts) - 1)] + [max_row]

            # Medium bottom border at each group boundary (except last group)
            for end_row in group_ends[:-1]:
                for col_idx in range(1, max_col + 1):
                    c = ws.cell(row=end_row, column=col_idx)
                    b = c.border
                    c.border = Border(left=b.left, right=b.right, top=b.top, bottom=medium)

            # Medium right border on group_col (all rows, including header)
            for row_idx in range(1, max_row + 1):
                c = ws.cell(row=row_idx, column=group_col)
                b = c.border
                c.border = Border(left=b.left, right=medium, top=b.top, bottom=b.bottom)

        if self.xlsx_path.exists():
            wb = load_workbook(self.xlsx_path)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        # ── Summary sheet ──────────────────────────────────────
        # Combine per-(model, dataset) rows + per-model "total" rows.
        # Same row structure as fwt_curves so per-dataset comparison plots
        # in Pass 3 can filter by dataset.
        all_new_rows = list(new_summaries) + list(model_total_summaries or [])
        new_pairs = {
            (str(s.get("model", "")), str(s.get("dataset", "")))
            for s in all_new_rows
        }

        existing_summary = []
        if "summary" in wb.sheetnames:
            ws_old = wb["summary"]
            old_headers = [cell.value for cell in ws_old[1]]
            raw_headers = [REVERSE_DISPLAY.get(h, h) for h in old_headers]
            has_dataset_col = "dataset" in raw_headers
            _last_model = None
            for row in ws_old.iter_rows(min_row=2, values_only=True):
                r = dict(zip(raw_headers, row))
                # Forward-fill merged model cells
                if r.get("model") is None:
                    r["model"] = _last_model
                else:
                    _last_model = r["model"]
                # Legacy migration: pre-fix summary rows were per-model totals
                if not has_dataset_col:
                    r["dataset"] = "total"
                m  = str(r.get("model", ""))
                ds = str(r.get("dataset", ""))
                if (m, ds) not in new_pairs:
                    existing_summary.append(r)
            del wb["summary"]

        all_total = existing_summary + all_new_rows
        all_total.sort(key=lambda r: (
            str(r.get("model", "")),
            "\xff" if str(r.get("dataset", "")) == "total" else str(r.get("dataset", ""))
        ))
        ws_sum = wb.create_sheet("summary", 0)
        if all_total:
            # Build headers: model, dataset first, then everything else (union over rows)
            seen_keys = []
            for r in all_total:
                for k in r.keys():
                    if k not in seen_keys:
                        seen_keys.append(k)
            headers = ["model", "dataset"] + [
                k for k in seen_keys
                if k not in ("model", "dataset") and k not in EXCLUDE_COLS
            ]
            ws_sum.append(headers)
            for row in all_total:
                ws_sum.append([row.get(h, None) for h in headers])
            _apply_sheet_style(ws_sum, headers)
            _merge_first_col(ws_sum)
            _apply_borders(ws_sum, group_col=1)

        # ── Per-model element sheets ───────────────────────────
        models_to_update = {s.get("model") for s in new_summaries if s.get("model")}
        for mlip_name in models_to_update:
            sheet_name = mlip_name[:31]

            # Preserve element rows for datasets not being updated in this run
            newly_computed = set(model_elem_data.get(mlip_name, {}).keys()) - {"total"}
            preserved_dfs = {}
            if sheet_name in wb.sheetnames:
                ws_old = wb[sheet_name]
                old_raw_headers = [REVERSE_DISPLAY.get(c.value, c.value) for c in ws_old[1]]
                if "dataset" in old_raw_headers:
                    rows_by_ds = {}
                    _last_ds = None
                    for row in ws_old.iter_rows(min_row=2, values_only=True):
                        r = dict(zip(old_raw_headers, row))
                        # Forward-fill merged cells (non-top cells read as None)
                        if r.get("dataset") is None:
                            r["dataset"] = _last_ds
                        else:
                            _last_ds = r["dataset"]
                        ds = str(r.get("dataset", ""))
                        if ds not in newly_computed and ds != "total":
                            rows_by_ds.setdefault(ds, []).append(r)
                    for ds_name, rows in rows_by_ds.items():
                        if rows:
                            preserved_dfs[ds_name] = pd.DataFrame(rows, columns=old_raw_headers)
                del wb[sheet_name]

            entry_dict = model_elem_data.get(mlip_name, {})
            # Merge preserved + newly computed keys and sort together (total always last)
            all_ds_keys = sorted(set(preserved_dfs.keys()) | {k for k in entry_dict if k != "total"})
            if "total" in entry_dict:
                all_ds_keys.append("total")

            dataset_dfs = []
            for dataset_key in all_ds_keys:
                if dataset_key in preserved_dfs:
                    dataset_dfs.append(preserved_dfs[dataset_key])
                else:
                    entry = entry_dict[dataset_key]
                    for type_label in ("normal", "anomaly"):
                        edf = entry.get(type_label)
                        if edf is None or edf.empty:
                            continue
                        edf = edf.copy()
                        edf.insert(0, "type", type_label)
                        edf.insert(0, "dataset", dataset_key)
                        dataset_dfs.append(edf)

            if not dataset_dfs:
                continue

            combined = pd.concat(dataset_dfs, ignore_index=True)
            ws_m = wb.create_sheet(sheet_name)
            elem_headers = list(combined.columns)
            ws_m.append(elem_headers)
            for _, row in combined.iterrows():
                ws_m.append([row[c] for c in elem_headers])
            _apply_sheet_style(ws_m, elem_headers)
            _merge_first_col(ws_m)
            _apply_borders(ws_m)

        # ── FwT curves sheet ───────────────────────────────────
        _FWT_KEY_T = [0.05, 0.10, 0.15, 0.20, 0.30, 0.50, 0.70, 1.00]
        _FWT_COLS  = [f"FwT@{t:.2f}" for t in _FWT_KEY_T]
        fwt_headers = ["model", "dataset"] + _FWT_COLS

        fwt_rows = []
        if "fwt_curves" in wb.sheetnames:
            ws_fwt_old = wb["fwt_curves"]
            old_fwt_headers = [cell.value for cell in ws_fwt_old[1]]
            # Normalize display-name headers back to raw keys ("Model" → "model" etc.)
            raw_fwt_headers = [REVERSE_DISPLAY.get(h, h) for h in old_fwt_headers]
            if raw_fwt_headers and "model" in raw_fwt_headers:
                _last_fwt_model = None
                for row in ws_fwt_old.iter_rows(min_row=2, values_only=True):
                    r = dict(zip(raw_fwt_headers, row))
                    # Forward-fill merged cells (non-top cells read as None)
                    if r.get("model") is None:
                        r["model"] = _last_fwt_model
                    else:
                        _last_fwt_model = r["model"]
                    model_val = str(r.get("model", ""))
                    if model_val not in models_to_update:
                        fwt_rows.append(r)
                    else:
                        # Preserve rows for datasets not being updated in this run
                        dataset_val = str(r.get("dataset", ""))
                        newly_fwt = set(model_fwt_data.get(model_val, {}).keys()) - {"total"}
                        if dataset_val not in newly_fwt and dataset_val != "total":
                            fwt_rows.append(r)
            del wb["fwt_curves"]

        for mlip_name in models_to_update:
            for dataset_key, (thresholds, fwt_pct) in model_fwt_data.get(mlip_name, {}).items():
                vals = np.interp(_FWT_KEY_T, thresholds, fwt_pct)
                row = {"model": mlip_name, "dataset": dataset_key}
                for col, val in zip(_FWT_COLS, vals):
                    row[col] = round(float(val), 2)
                fwt_rows.append(row)

        if fwt_rows:
            # Sort by model then dataset (total always last within each model)
            fwt_rows.sort(key=lambda r: (
                str(r.get("model", "")),
                "\xff" if str(r.get("dataset", "")) == "total" else str(r.get("dataset", ""))
            ))
            ws_fwt = wb.create_sheet("fwt_curves")
            ws_fwt.append(fwt_headers)
            for row in fwt_rows:
                ws_fwt.append([row.get(h) for h in fwt_headers])

            for col_idx, h in enumerate(fwt_headers, 1):
                cell = ws_fwt.cell(row=1, column=col_idx)
                cell.value     = HEADER_DISPLAY.get(h, h)
                cell.font      = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            ws_fwt.row_dimensions[1].height = 20

            _autosize_columns(ws_fwt, fwt_headers)

            for row_idx, row in enumerate(ws_fwt.iter_rows(min_row=2), 2):
                ws_fwt.row_dimensions[row_idx].height = 25
                for col_idx, cell in enumerate(row, 1):
                    h = fwt_headers[col_idx - 1]
                    cell.alignment = Alignment(
                        horizontal="left" if h == "model" else "center",
                        vertical="center"
                    )

                    sanitized = _sanitize_value(cell.value, h)
                    if sanitized == "-" and cell.value != "-":
                        cell.value = "-"
                        cell.number_format = "@"
                        continue

                    if h not in ("model", "dataset"):
                        cell.number_format = "0.00"
            _merge_first_col(ws_fwt)
            _apply_borders(ws_fwt)

        wb.save(self.xlsx_path)
        print(f"  results.xlsx saved: {self.xlsx_path}")

    # ── Plots (per model) ─────────────────────────────────────

    def _plot_model(self, mlip_name, elem_df, records_df, out_dir: Path,
                    skip_heatmap: bool = False):
        try:
            if records_df is None or records_df.empty:
                return

            df   = records_df.copy()
            edir = out_dir / "energy"
            fdir = out_dir / "force"
            edir.mkdir(parents=True, exist_ok=True)
            fdir.mkdir(parents=True, exist_ok=True)

            plot_parity_elem(
                df["e_form_dft"].values,
                df["e_form_mlip"].values,
                df["element"].values,
                title=f"{mlip_name} — $\\mathbf{{E_{{form}}}}$ parity",
                xlabel="DFT $\\mathbf{E_{form}}$ (eV/atom)",
                ylabel="MLIP $\\mathbf{E_{form}}$ (eV/atom)",
                save_path=edir / "Eform.png",
                dpi=self.dpi,
                mlip_name=mlip_name,
            )

            f_dft_rows, f_mlip_rows, f_elem_rows = [], [], []
            for _, row in df.iterrows():
                fd = np.asarray(row["dft_forces"], dtype=float)
                fm = np.asarray(row["mlip_forces"], dtype=float)
                if fd.shape != fm.shape or fd.ndim != 2:
                    continue
                n_at = fd.shape[0]
                f_dft_rows.append(fd.reshape(-1))
                f_mlip_rows.append(fm.reshape(-1))
                f_elem_rows.extend([row["element"]] * (n_at * 3))

            if f_dft_rows:
                plot_parity_elem(
                    np.concatenate(f_dft_rows),
                    np.concatenate(f_mlip_rows),
                    np.array(f_elem_rows, dtype=str),
                    title=f"{mlip_name} — $\\mathbf{{Force}}$ parity",
                    xlabel="DFT $\\mathbf{Force}$ (eV/Å)",
                    ylabel="MLIP $\\mathbf{Force}$ (eV/Å)",
                    save_path=fdir / "Force.png",
                    dpi=self.dpi,
                    iqr_filter=False,
                    mlip_name=mlip_name,
                    unit="eV/Å",
                )

            if not skip_heatmap and elem_df is not None and not elem_df.empty:
                single_elem_df = elem_df[elem_df["element"].apply(_is_single_element)]
                if not single_elem_df.empty:
                    count_dict = dict(zip(single_elem_df["element"], single_elem_df["n_samples"]))
                    for col, fname, higher_is_better, dest in [
                        ("E_form_MAE",      "heatmap_MAE.png",      False, edir),
                        ("E_form_RMSE",     "heatmap_RMSE.png",     False, edir),
                        ("E_form_R2",       "heatmap_R2.png",       True,  edir),
                        ("E_form_Pearson",  "heatmap_Pearson.png",  True,  edir),
                        ("E_form_Spearman", "heatmap_Spearman.png", True,  edir),
                        ("force_MAE",       "heatmap_MAE.png",      False, fdir),
                        ("force_RMSE",      "heatmap_RMSE.png",     False, fdir),
                        ("force_R2",        "heatmap_R2.png",       True,  fdir),
                        ("force_Pearson",   "heatmap_Pearson.png",  True,  fdir),
                        ("force_Spearman",  "heatmap_Spearman.png", True,  fdir),
                        ("force_cosine",    "heatmap_cosine.png",   True,  fdir),
                    ]:
                        if col not in single_elem_df.columns:
                            continue
                        metric_dict = dict(zip(single_elem_df["element"], single_elem_df[col]))
                        plot_periodic_heatmap(
                            metric_dict,
                            title=f"{mlip_name} — {col}",
                            save_path=dest / fname,
                            higher_is_better=higher_is_better,
                            dpi=self.dpi,
                            count_dict=count_dict,
                        )

            print(f"  plots saved: {out_dir}")

        except Exception as exc:
            print(f"  [WARN] {mlip_name} plot failed: {exc}")

    # ── Plots (global total) from xlsx ────────────────────────

    def _plot_global_total_from_xlsx(self, dataset_names: list):
        """
        Generate analysis/total/{dataset}/ and analysis/total/total/
        entirely from xlsx — no DB reading.
        """
        try:
            if not self.xlsx_path.exists():
                print("  [WARN] xlsx not found, skipping total plots.")
                return

            summary_df = _normalize_xlsx_df(
                pd.read_excel(self.xlsx_path, sheet_name="summary")
            )
            if summary_df.empty or "model" not in summary_df.columns:
                return
            # merged cells leave Model as NaN on secondary rows — restore
            summary_df["model"] = summary_df["model"].ffill()

            fwt_df = pd.DataFrame()
            xl = pd.ExcelFile(self.xlsx_path)
            if "fwt_curves" in xl.sheet_names:
                fwt_df = _normalize_xlsx_df(
                    pd.read_excel(xl, sheet_name="fwt_curves")
                )
                # merged cells in the xlsx leave Model as NaN on secondary
                # rows (e.g. the "total" row of each model) — restore it
                if "model" in fwt_df.columns:
                    fwt_df["model"] = fwt_df["model"].ffill()

            model_names = summary_df["model"].astype(str).unique().tolist()

            for dataset_name in dataset_names:
                out_dir = self.analysis_root / "total" / dataset_name
                out_dir.mkdir(parents=True, exist_ok=True)
                self._plot_comparison_from_xlsx(
                    summary_df, fwt_df, model_names, dataset_name, out_dir
                )

            total_out = self.analysis_root / "total" / "total"
            total_out.mkdir(parents=True, exist_ok=True)
            self._plot_comparison_from_xlsx(
                summary_df, fwt_df, model_names, None, total_out
            )

            print(f"  global total plots saved: {self.analysis_root / 'total'}")

        except Exception as exc:
            print(f"  [WARN] global total plot failed: {exc}")

    def _plot_comparison_from_xlsx(self, summary_df, fwt_df, model_names,
                                   dataset_filter, out_dir: Path):
        """
        Bar/pareto/FwT comparison plots across models, reading from xlsx only.
        summary_df: the 'summary' sheet — (model, dataset) rows + total per model
        fwt_df    : the 'fwt_curves' sheet — same row structure
        dataset_filter: dataset name, or None for the per-model total comparison
        """
        try:
            df = summary_df.copy()
            df = df.replace("-", np.nan)
            df["model"] = df["model"].astype(str)

            dataset_key = dataset_filter or "total"

            # Filter summary rows by dataset. Legacy summaries lacking the
            # dataset column can only support the "total" view.
            if "dataset" in df.columns:
                df["dataset"] = df["dataset"].astype(str)
                df = df[df["dataset"] == dataset_key]
            elif dataset_key != "total":
                return

            df = df.drop_duplicates(subset=["model"], keep="last").reset_index(drop=True)
            if df.empty:
                return

            # Models that actually have data for this dataset
            ds_model_names = df["model"].unique().tolist()

            if not fwt_df.empty:
                fwt_df = fwt_df.replace("-", np.nan)

            time_col = "cluster_time_med_s"
            time_per_step = (
                df.set_index("model")[time_col].reindex(ds_model_names).astype(float).tolist()
                if time_col in df.columns else [float("nan")] * len(ds_model_names)
            )

            if "n_normal" in df.columns and "n_anomaly" in df.columns:
                idx = df.set_index("model")
                nn = idx["n_normal"].reindex(ds_model_names).astype(float).values
                na = idx["n_anomaly"].reindex(ds_model_names).astype(float).values
                normal_rate = (nn / (nn + na + 1e-12)).tolist()
            else:
                normal_rate = [float("nan")] * len(ds_model_names)

            metrics = [
                ("E_form_MAE",      "E_form MAE (eV/atom)"),
                ("E_form_RMSE",     "E_form RMSE (eV/atom)"),
                ("E_form_R2",       "E_form $R^2$"),
                ("E_form_Pearson",  "E_form Pearson $r$"),
                ("E_form_Spearman", "E_form Spearman $\\rho$"),
                ("force_MAE",       "Force MAE (eV/Å)"),
                ("force_RMSE",      "Force RMSE (eV/Å)"),
                ("force_R2",        "Force $R^2$"),
                ("force_Pearson",   "Force Pearson $r$"),
                ("force_Spearman",  "Force Spearman $\\rho$"),
                ("force_cosine",    "Force cosine similarity"),
            ]

            subset_dir = out_dir
            subset_dir.mkdir(parents=True, exist_ok=True)

            for col, ylabel in metrics:
                if col not in df.columns:
                    continue
                idx = df.set_index("model")
                values = idx[col].reindex(ds_model_names).astype(float).tolist()
                ascending = not any(kw in col for kw in ("R2", "Pearson", "Spearman", "cosine"))
                plot_total_bar(
                    model_names=ds_model_names,
                    values=values,
                    metric_name=f"{dataset_key}_{col}",
                    ylabel=ylabel,
                    save_path=subset_dir / f"{col}.png",
                    dpi=self.dpi,
                    ascending=ascending,
                )

            acc_vals = df.set_index("model")["E_form_MAE"].reindex(ds_model_names).astype(float).tolist() \
                if "E_form_MAE" in df.columns else [float("nan")] * len(ds_model_names)

            plot_pareto(
                model_names=ds_model_names,
                x_vals=time_per_step,
                y_vals=acc_vals,
                title=f"{dataset_key} —Accuracy-Efficiency",
                xlabel="Time per step (s)",
                ylabel="E_form MAE (eV/atom)",
                save_path=subset_dir / "pareto_accuracy_efficiency.png",
                dpi=self.dpi,
                maximize_y=False,
            )
            plot_pareto(
                model_names=ds_model_names,
                x_vals=time_per_step,
                y_vals=normal_rate,
                title=f"{dataset_key} —Robustness-Efficiency",
                xlabel="Time per step (s)",
                ylabel="Normal rate",
                save_path=subset_dir / "pareto_robustness_efficiency.png",
                dpi=self.dpi,
                maximize_y=True,
            )

            # FwT curve comparison from fwt_curves sheet
            if not fwt_df.empty:
                fwt_sub = fwt_df[fwt_df["dataset"].astype(str) == dataset_key]
                fwt_cols = [c for c in fwt_df.columns if c.startswith("FwT@")]
                thresholds_key = np.array([float(c.replace("FwT@", "")) for c in fwt_cols])

                model_curves = {}
                for _, row in fwt_sub.iterrows():
                    mn = str(row.get("model", ""))
                    if mn not in ds_model_names:
                        continue
                    fwt_vals = np.array([float(row[c]) for c in fwt_cols])
                    model_curves[mn] = (thresholds_key, fwt_vals)

                if model_curves:
                    plot_fwt_curve(
                        model_curves,
                        save_path=out_dir / "fwt_curve.png",
                        dpi=self.dpi,
                    )

        except Exception as exc:
            print(f"  [WARN] comparison plot failed ({dataset_filter or 'total'}): {exc}")